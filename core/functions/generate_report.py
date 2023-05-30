import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime, timedelta
from core.database.operations import DataBase

db = DataBase()

STATUS = {
    'sent': 'Ожидает реагирования',
    None: 'Ожидает реагирования',
    'work': 'В работе',
    'done': 'Отработано',
    'reject': 'Отклонено'
  }
PUBLIC_USERS_WIDTH = [10, 25, 20, 13, 35, 20, 20, 20, 13, 20]
PUBLIC_USERS_NOT_SEND_WIDTH = [10, 25, 20, 13, 35, 20, 20, 20, 13, 20]
APPEALS_WIDTH = [10, 25, 20, 13, 35, 20, 20, 20, 13, 20, 20, 30]

PUBLIC_USERS_HEADER = ['ID', 'Имя', 'username', 'telegram ID', 'Адрес', 'Вид животного', 'Количество животных',
                       'Контакты', 'Медиа', 'Дата подписки']
PUBLIC_USERS_NOT_SEND_HEADER = ['ID', 'Имя', 'username', 'telegram ID', 'Адрес', 'Вид животного', 'Количество животных',
                                'Контакты', 'Медиа', 'Дата подписки']
APPEALS_HEADER = ['ID', 'Имя', 'username', 'telegram ID', 'Адрес', 'Вид животного', 'Количество животных',
                  'Контакты', 'Медиа', 'Дата обращения', 'Статус', 'Ответ оператора']


async def cell_edit(sheet: Worksheet, value,
                    range_: str,
                    bold: bool = False,
                    size: int = 12,
                    color: str = '000000',
                    name: str = 'Times New Roman',
                    alignment: Alignment = None):
    sheet[range_] = value
    sheet[range_].font = Font(bold=bold, size=size, name=name, color=color)
    sheet[range_].alignment = alignment


async def column_dimensions(sheet: Worksheet, width_columns: list):
    for i, width in enumerate(width_columns):
        column = get_column_letter(i+1)
        sheet.column_dimensions[column].width = width


async def create_header_table(sheet: Worksheet, columns: list):
    for i, column in enumerate(columns):
        await cell_edit(sheet, column, f"{get_column_letter(i+1)}1", bold=True,
                        alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))


async def insert_data_public_users(sheet: Worksheet, data: pd.DataFrame):
    for i, row in data.iterrows():
        await cell_edit(sheet, row['user_id'], f"A{i + 2}")
        await cell_edit(sheet, row['fullname'], f"B{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['username'], f"C{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['telegram_id'], f"D{i + 2}")
        await cell_edit(sheet, row['address'], f"E{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['animal'], f"F{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['count_animal'], f"G{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['contact'], f"H{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['media'], f"I{i + 2}")
        await cell_edit(sheet, row['date_sub'], f"J{i + 2}")


async def insert_data_appeals(sheet: Worksheet, data: pd.DataFrame):
    count_ban = 0
    count_reject = 0
    count_work = 0
    count_done = 0
    count_send = 0
    for i, row in data.iterrows():
        await cell_edit(sheet, row['appeal_id'], f"A{i + 2}")
        await cell_edit(sheet, row['fullname'], f"B{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['username'], f"C{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['telegram_id'], f"D{i + 2}")
        await cell_edit(sheet, row['address'], f"E{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['animal'], f"F{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['count_animal'], f"G{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['contact'], f"H{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['media'], f"I{i + 2}")
        await cell_edit(sheet, row['date_create'], f"J{i + 2}")

        try:
            none_ban = np.isnan(row['ban'])
        except:
            none_ban = False if row['ban'] else True

        if not none_ban:
            status = 'Пользователь заблокирован'
            count_ban += 1
        else:
            if row['status'] == 'done':
                count_done += 1
            elif row['status'] == 'work':
                count_work += 1
            elif row['status'] == 'reject':
                count_reject += 1
            elif row['status'] == 'sent' or row['status'] is None:
                count_send += 1
            status = STATUS[row['status']]
        await cell_edit(sheet, status, f"K{i + 2}", alignment=Alignment(wrap_text=True))
        await cell_edit(sheet, row['answer'], f"L{i + 2}", alignment=Alignment(wrap_text=True))

    return count_done, count_work, count_send, count_reject, count_ban


async def generate_report():
    date_now = datetime.now().replace(hour=9, minute=0, second=0, microsecond=0)
    date_previous = date_now - timedelta(days=1)
    public_users = await db.get_statistic_public_users(date_start=date_previous, date_end=date_now)
    appeals = await db.get_statistic_appeals(date_start=date_previous, date_end=date_now)
    number_public_users = await db.get_number_public_users(date_end=date_now)

    df_public_users = pd.DataFrame(public_users)
    df_appeals = pd.DataFrame(appeals)

    if not df_appeals.empty:
        # Датафрейм с начавшими заполнять, но не отправили
        df_not_send = pd.merge(df_public_users, df_appeals, how='left', on='telegram_id', suffixes=('', '_y'))
        df_not_send = df_not_send[(~df_not_send['address'].isna()) & (df_not_send['status'].isna())].reset_index()
    else:
        df_not_send = pd.DataFrame()
    book = openpyxl.Workbook()
    sheet_public_users = book.create_sheet('Подписчики')
    sheet_appeals = book.create_sheet('Обращения')
    sheet_not_send = book.create_sheet('Не отправленные')
    sheet = book.active
    book.remove_sheet(worksheet=sheet)

    await column_dimensions(sheet_public_users, PUBLIC_USERS_WIDTH)
    await column_dimensions(sheet_appeals, APPEALS_WIDTH)
    await column_dimensions(sheet_not_send, PUBLIC_USERS_NOT_SEND_WIDTH)

    await create_header_table(sheet_public_users, PUBLIC_USERS_HEADER)
    await create_header_table(sheet_appeals, APPEALS_HEADER)
    await create_header_table(sheet_not_send, PUBLIC_USERS_NOT_SEND_HEADER)

    await insert_data_public_users(sheet_public_users, df_public_users)
    number_done, number_work, number_send, number_reject, number_ban = \
        await insert_data_appeals(sheet_appeals, df_appeals)
    await insert_data_public_users(sheet_not_send, df_not_send)

    path_report = f"data/Статистика.xlsx"
    book.save(path_report)
    book.close()

    text_statistic = f"Бот «Спаси сову 🦉»: <b>{date_previous.strftime('%d.%m.%Y %H:%M')} - " \
                     f"{date_now.strftime('%d.%m.%Y %H:%M')}</b>\n\n" \
                     f"Количество подписчиков: <b>{number_public_users}</b> (+{len(public_users)} человек)\n" \
                     f"Количество начавших заполнять бота, но не отправили: {len(df_not_send.index)}\n" \
                     f"Количество обращений поступивших операторам - {len(appeals)}:\n" \
                     f" - Отработано: <b>{number_done}</b>\n" \
                     f" - В работе: <b>{number_work}</b>\n" \
                     f" - В ожидании реагирования: <b>{number_send}</b>\n" \
                     f" - Отклонено: <b>{number_reject}</b>\n" \
                     f" - Заблокировано: <b>{number_ban}</b>\n\n" \
                     f"📎 Подробная информация в прикрепленном файле."

    return text_statistic, path_report
